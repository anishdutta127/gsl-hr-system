"""
Import Employee_Muster.xlsx into src/data/employees.json.

Seed-style: writes the final JSON directly (commits via git, not queue).
The runtime queue pattern is for user-originated writes; 126 pre-existing
employees is initial state that lives in the repo.

Two-pass:
  1. Parse rows -> Employee records with reportingManagerName string.
  2. Resolve reportingManagerId by matching manager name to Employee.name.
     Unmatched rows (e.g., CEO reports to 'PHM') leave id null and log.

Rows with no Official Email ID are still imported as Employee records
(they can't log in but they do exist on the muster).
"""

from __future__ import annotations

import hashlib
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


MUSTER_PATH = Path("onedrive-data/seed/Employee_Muster.xlsx")
OUT_PATH = Path("src/data/employees.json")
LOG_PATH = Path("logs/muster_import.log")


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z")


def stable_uuid(employee_code: str) -> str:
    """Deterministic UUIDv5-ish id from employee code. Re-importing the muster
    produces the same id, so an apply-queue cycle does not duplicate rows."""
    h = hashlib.sha1(f"gsl-employee:{employee_code}".encode("utf-8")).hexdigest()
    return f"{h[0:8]}-{h[8:12]}-5{h[13:16]}-{h[16:20]}-{h[20:32]}"


def iso_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value).date().isoformat()
        except ValueError:
            return value.strip() or None
    return str(value)


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_phone(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = re.sub(r"[^\d+]", "", s)
    if s and not s.startswith("+") and len(s) == 10:
        return f"+91-{s}"
    return s


def parse_muster(path: Path) -> list[dict]:
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    now = now_iso()
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]

    employees: list[dict] = []
    missing_email: list[str] = []

    for idx, row in enumerate(data_rows, start=2):
        employee_code = clean(row[0])
        if not employee_code:
            continue

        title = clean(row[1])
        name = clean(row[2])
        doj = iso_date(row[3])
        tenure = row[6]
        designation = clean(row[7])
        department = clean(row[8])
        reporting_manager = clean(row[9])
        confirm_date = iso_date(row[10])
        location = clean(row[11]) or "Mumbai"
        official_email = clean(row[12]).lower()
        gender = clean(row[13])
        dob = iso_date(row[14])
        age = row[15]
        marital_status = clean(row[16])
        phone = clean_phone(row[17])
        address = clean(row[18])
        personal_email = clean(row[19]).lower()

        if not official_email:
            missing_email.append(f"row {idx}: {name} ({employee_code})")

        employees.append({
            "id": stable_uuid(employee_code),
            "employeeCode": employee_code,
            "title": title or None,
            "name": name,
            "email": official_email or personal_email,
            "phone": phone or None,
            "designation": designation,
            "department": department,
            "reportingTo": reporting_manager or None,
            "reportingManagerId": None,  # filled in pass 2
            "location": location,
            "dateOfJoining": doj,
            "status": "Active",
            "confirmationDate": confirm_date,
            "tenureYears": tenure if isinstance(tenure, (int, float)) else None,
            "dateOfBirth": dob,
            "age": age if isinstance(age, (int, float)) else None,
            "gender": gender or None,
            "maritalStatus": marital_status or None,
            "address": address or None,
            "personalEmail": personal_email or None,
            "officialEmailMissing": not bool(official_email),
            "createdAt": now,
            "createdBy": "muster-import",
            "auditLog": [{
                "timestamp": now,
                "user": "muster-import",
                "action": "employee.create",
                "after": {
                    "employeeCode": employee_code,
                    "name": name,
                    "designation": designation,
                    "department": department,
                },
                "notes": "Imported from Employee_Muster.xlsx",
            }],
        })

    return employees, missing_email


def resolve_managers(employees: list[dict]) -> list[str]:
    """Match reportingTo name against employees.name with a first-name fallback.
    The muster uses casual names like 'Balu R' / 'Shubhangi' for references but
    full names for rows; a first-name-exact match catches those cases."""
    by_name: dict[str, str] = {}
    by_first: dict[str, list[str]] = {}
    for e in employees:
        full = e["name"].strip().lower()
        by_name[full] = e["id"]
        first = full.split()[0] if full else ""
        if first:
            by_first.setdefault(first, []).append(e["id"])

    unresolved: list[str] = []
    for e in employees:
        raw = e.get("reportingTo")
        if not raw:
            continue
        key = raw.strip().lower()
        match_id = by_name.get(key)
        if not match_id:
            first = key.split()[0] if key else ""
            candidates = by_first.get(first, [])
            if len(candidates) == 1:
                match_id = candidates[0]
        if match_id:
            e["reportingManagerId"] = match_id
        else:
            unresolved.append(f"{e['name']} ({e['employeeCode']}) reports to '{raw}' -> no match")
    return unresolved


def main() -> int:
    if not MUSTER_PATH.exists():
        print(f"Not found: {MUSTER_PATH}", file=sys.stderr)
        return 1

    employees, missing_email = parse_muster(MUSTER_PATH)
    unresolved = resolve_managers(employees)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(employees, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with LOG_PATH.open("w", encoding="utf-8") as f:
        f.write(f"Muster import run at {now_iso()}\n")
        f.write(f"Employees imported: {len(employees)}\n")
        f.write(f"Missing official email ({len(missing_email)}):\n")
        for line in missing_email:
            f.write(f"  {line}\n")
        f.write(f"\nUnresolved reporting managers ({len(unresolved)}):\n")
        for line in unresolved:
            f.write(f"  {line}\n")

    print(f"Wrote {len(employees)} employees to {OUT_PATH}")
    print(f"Missing official email: {len(missing_email)}")
    print(f"Unresolved reporting managers: {len(unresolved)}")
    print(f"Details: {LOG_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
